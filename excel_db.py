# excel_db.py
import openpyxl

category = {'RFP':3,'RFI':3,'GVE Support':1,'Complex Design':2}
complexity ={'High':3,'Medium':2,'Low':1}
status = {'WiP':1,'CE Pending':.1, 'Close':0}
headers = ['id','customer','project','category','cat_effort','complexity','com_effort','effort','status','status_effort','region','tsa','user','total_hrs','webex_user']
webex_user = {'AH':'alejher2@cisco.com', 'FQ':'frquiroz@cisco.com'}

excel_file_path1 = 'gve_sp_americas_report.xlsx'
excel_file_path2 = '/mnt/c/Users/frquiroz/OneDrive - Cisco/Documents/00 GVE/gve fy24/GVE SP Americas Report.xlsx'
gve_sheet = 'Bot'

class gve_record:
    def __init__(self, headers):
        # Initialize a dictionary with keys from the headers list and None as values
        self.headers = headers
        self.records = {header: None for header in headers}
    
    def ask_for_new_record(self):
        # Ask the user for input for each header and create a new record
        new_record = {}
        for header in self.headers:
            if 'effort' not in header:
                self.records[header] = input(f"Enter {header}: ")
            else:
                self.records[header] = None

        return self.records
    
    # no es util
    def add_new_record(self, record):
        # Adds the new record to the records dictionary
        for key, value in record.items():
            if key in self.records:
                self.records[key] = value
            else:
                print(f"Record with key '{key}' does not exist in the headers.")
    
    # no es util
    def update_record(self, key, value):
        # Update an existing record
        if key in self.records:
            self.records[key] = value
            return True
        else:
            print(f"Record with key '{key}' does not exist.")
            return False
    
    def get_records(self):
        # Return all records
        return self.records
    
    def get_headers(self):
        return self.headers
sample = {'data': 'add', 'category': 'RFP', 'complexity': 'High', 'customer': '2', 'project': '3', 'status': 'WiP'}

#new_gve_record = gve_record(headers)
def add_gve_record(record_input):
    
    new_gve_record = gve_record(headers)
    
    for key, value in record_input.items():
        if key in new_gve_record.records.keys():
            new_gve_record.records[key] = record_input[key]
    new_gve_record.records['cat_effort'] = category[new_gve_record.records['category']]
    new_gve_record.records['com_effort'] = complexity[new_gve_record.records['complexity']]
    new_gve_record.records['com_effort'] = complexity[new_gve_record.records['complexity']]
    new_gve_record.records['status_effort'] = status[new_gve_record.records['status']]
    new_gve_record.records['webex_user'] = webex_user[new_gve_record.records['tsa']]
    new_gve_record.records['effort'] = new_gve_record.records['cat_effort'] * new_gve_record.records['com_effort']
    new_gve_record.records['total_hrs'] = new_gve_record.records['effort'] * new_gve_record.records['status_effort']
    a, b = get_last_and_next_id()

    new_gve_record.records['id'] = b
    
    return update_excel(new_gve_record.records)


def update_excel(new_record):
    print(f"El nuevo registro es {new_record}")
    new_data = list(new_record.values())
    
    workbook = openpyxl.load_workbook(excel_file_path2)

    # Select a worksheet by name
    worksheet = workbook[gve_sheet]

    worksheet.append(new_data)
    
    workbook.save(filename=excel_file_path2)

    workbook.close()
    
    for row in worksheet.iter_rows(values_only=True):
        print(row)

    return f'ok se actualizo excel {new_data}'


def find_matching_rows_by_email(email_to_match, email_column_name='webex_user', 
                                include_columns=['ID','Customer', 'Project', 'Category','Complexity','Status']):
    """
    Finds rows in an Excel file where the specified email column matches the given email address 
    and returns selected columns: Customer, Project, and Status.

    :param excel_file_path: Path to the Excel file
    :param email_to_match: Email address to match in the column
    :param email_column_name: Column header name that contains the email addresses (default is 'webex_user')
    :param include_columns: List of columns to include in the result (default includes Customer, Project, and Status)
    :return: List of dictionaries with the selected columns from rows that match the email address
    """
    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(excel_file_path2)
    ws = wb.active

    # Find the column indices for the email column and include_columns
    columns_indices = {}
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == email_column_name:
            columns_indices[email_column_name] = col[0].column - 1  # Adjust for zero-indexing
        elif col[0].value in include_columns:
            columns_indices[col[0].value] = col[0].column - 1  # Adjust for zero-indexing

    # Check if all required columns were found
    if email_column_name not in columns_indices:
        print(f"Column '{email_column_name}' not found.")
        return []
    for col_name in include_columns:
        if col_name not in columns_indices:
            print(f"Column '{col_name}' not found.")
            return []

    # Iterate over rows and check if the email matches
    matching_rows = []
    for row in ws.iter_rows(min_row=2):  # Assuming the first row is the header
        if row[columns_indices[email_column_name]].value == email_to_match:
            # Extract the selected columns
            row_data = {col_name: row[idx].value for col_name, idx in columns_indices.items() if col_name in include_columns}
            matching_rows.append(row_data)

    # Close the workbook if you're done with it
    wb.close()

    return matching_rows



def get_last_and_next_id(id_column_name='ID'):
    """
    Gets the last and next ID from a specified column in an Excel file.

    :param excel_file_path: Path to the Excel file
    :param id_column_name: Column header name that contains the ID records (default is 'ID')
    :return: A tuple containing the last ID and the next ID
    """
    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(excel_file_path2)
    ws = wb.active

    # Find the column index for the ID column
    id_column_index = None
    for cell in ws[1]:  # Assuming the first row is the header
        if cell.value == id_column_name:
            id_column_index = cell.column  # column index
            break

    if id_column_index is None:
        print(f"Column '{id_column_name}' not found.")
        return None, None

    # Find the last ID in the column
    last_id = None
    for cell in ws.iter_cols(min_col=id_column_index, max_col=id_column_index, min_row=2):
        for c in cell:
            if c.value is not None:
                last_id = c.value

    # Close the workbook
    wb.close()

    if last_id is None:
        print(f"No ID records found in column '{id_column_name}'.")
        return None, None

    # Assuming that the ID is numeric and the next ID is just an increment of the last ID
    try:
        last_id = int(last_id)  # Convert last ID to integer if it's not already
        next_id = last_id + 1
    except ValueError:
        print(f"Last ID '{last_id}' is not numeric.")
        return None, None

    return last_id, next_id
